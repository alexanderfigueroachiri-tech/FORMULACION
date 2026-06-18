"""Analyze Excel workbooks: formulas, cell refs, economic concepts."""
import json
import re
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent.parent / "BASE"
OUT = Path(__file__).resolve().parent.parent / "data" / "workbook_analysis.json"

CELL_REF = re.compile(
    r"(?:(?:'([^']+)'|([A-Za-z0-9_]+))!)?(\$?[A-Z]+\$?\d+)"
)

CONCEPT_MAP = {
    "ingreso": "Ingresos del proyecto",
    "gasto": "Costos operativos",
    "dep": "Depreciacion (gasto no desembolsable)",
    "uaii": "Utilidad antes de intereses e impuestos",
    "uai": "Utilidad antes de impuestos",
    "interes": "Intereses sobre deuda",
    "impuesto": "Impuesto sobre utilidades",
    "flujo": "Flujo de fondos",
    "prestamo": "Desembolso o amortizacion de prestamo",
    "npv": "Valor Actual Neto (VAN)",
    "irr": "Tasa Interna de Retorno (TIR)",
    "pmt": "Cuota constante de prestamo",
    "rf": "Tasa libre de riesgo",
    "beta": "Beta del activo",
    "riesgo": "Prima de riesgo o riesgo pais",
    "amortiz": "Amortizacion de capital",
    "capital": "Saldo de capital de deuda",
    "cuota": "Pago total del periodo",
    "van": "Valor Actual Neto",
    "tir": "Tasa Interna de Retorno",
    "cok": "Costo de oportunidad del capital",
    "rentab": "Rentabilidad del proyecto",
    "vaf": "Valor actual de flujos",
}


def guess_concept(label: str, formula: str) -> str | None:
    text = f"{label} {formula}".lower()
    for key, desc in CONCEPT_MAP.items():
        if key in text:
            return desc
    return None


def extract_refs(formula: str) -> list[dict]:
    if not formula or not str(formula).startswith("="):
        return []
    refs = []
    for match in CELL_REF.finditer(str(formula)):
        sheet = match.group(1) or match.group(2)
        cell = match.group(3).replace("$", "")
        refs.append({"sheet": sheet, "cell": cell})
    return refs


def row_label(ws, row_num: int) -> str:
    for col in range(1, 4):
        val = ws.cell(row=row_num, column=col).value
        if val and isinstance(val, str) and not val.startswith("="):
            return val.strip()
    return ""


def analyze_xlsx(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=False)
    result = {"file": path.name, "sheets": {}}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cells = []
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if val is None:
                    continue
                is_formula = isinstance(val, str) and val.startswith("=")
                if not is_formula and cell.data_type != "f":
                    continue
                formula = str(val)
                label = row_label(ws, cell.row)
                concept = guess_concept(label, formula)
                refs = extract_refs(formula)
                cells.append(
                    {
                        "cell": cell.coordinate,
                        "formula": formula,
                        "label": label,
                        "concept": concept,
                        "refs": refs[:10],
                        "is_key": any(
                            token in formula.upper()
                            for token in ("NPV", "IRR", "PMT", "SUM")
                        ),
                    }
                )
        if cells:
            result["sheets"][sheet_name] = {
                "formula_count": len(cells),
                "key_formulas": [c for c in cells if c["is_key"]][:30],
                "concepts": sorted({c["concept"] for c in cells if c["concept"]}),
                "chains": cells[:50],
            }
    wb.close()
    return result


def build_curriculum(analysis: list[dict]) -> dict:
    """Progressive learning path aligned with BEG06 materials."""
    return {
        "course": "BEG06 - Formulacion y Evaluacion de Proyectos",
        "methodology": "Aprendizaje progresivo: concepto -> tip -> practica Excel -> validacion",
        "estimated_hours": 4,
        "modules": [
            {
                "id": "m1-depreciacion",
                "order": 1,
                "title": "Depreciacion",
                "duration_min": 35,
                "source": [
                    "Capitulo V. Depreciacion Parte I.pdf",
                    "Capitulo V. Depreciacion Parte II.pdf",
                ],
                "objectives": [
                    "Entender depreciacion lineal y su rol en el flujo",
                    "Distinguir gasto contable vs desembolso",
                ],
                "tips": [
                    "La depreciacion reduce utilidad pero NO sale caja: se suma de vuelta al flujo.",
                    "Vida util y valor residual definen la cuota anual de depreciacion.",
                ],
                "steps": [
                    {
                        "type": "concept",
                        "title": "Que es la depreciacion",
                        "body": "Distribuye el costo de un activo en su vida util. Es egreso contable, no desembolso.",
                    },
                    {
                        "type": "formula",
                        "title": "Depreciacion lineal",
                        "formula": "Dep anual = (Inversion - Valor residual) / Vida util",
                        "excel_equiv": "=+(Inversion-VR)/Vida",
                    },
                    {
                        "type": "practice",
                        "worksheet": "depreciation_basic",
                        "title": "Calcula depreciacion",
                    },
                ],
            },
            {
                "id": "m2-fce",
                "order": 2,
                "title": "Flujo de Caja Economico (FCE)",
                "duration_min": 45,
                "source": ["S6 Flujo basico FCE y FCF.xlsx", "Capitulo VI Flujo de caja Parte 1.pptx"],
                "objectives": [
                    "Construir FCE periodo a periodo",
                    "Calcular UAII, UAI, impuestos y flujo de fondos economico",
                ],
                "tips": [
                    "UAII = Ingresos - Gastos - Depreciacion",
                    "Flujo economico = Neto + Depreciacion (sin financiamiento)",
                ],
                "steps": [
                    {
                        "type": "concept",
                        "title": "Cadena del FCE",
                        "body": "Ingresos -> Gastos -> Dep -> UAII -> Impuestos -> Neto -> +Dep -> Flujo fondos",
                    },
                    {
                        "type": "practice",
                        "worksheet": "fce_basic",
                        "title": "Completa el FCE de 3 periodos",
                        "source_file": "S6 Flujo basico FCE y FCF.xlsx",
                        "source_sheet": "FCE",
                    },
                ],
            },
            {
                "id": "m3-tasas",
                "order": 3,
                "title": "Tasa de actualizacion (CAPM + riesgo pais)",
                "duration_min": 30,
                "source": ["S6 Flujo basico FCE y FCF.xlsx"],
                "objectives": [
                    "Calcular tasa de descuento con CAPM",
                    "Incorporar riesgo pais al costo de capital",
                ],
                "tips": [
                    "Ke = Rf + Beta * (Rm - Rf) + Riesgo pais",
                    "Esta tasa es la que usas en NPV para VAN economico",
                ],
                "steps": [
                    {
                        "type": "formula",
                        "title": "CAPM extendido",
                        "formula": "Ke = Rf + Beta * (Rm - Rf) + Riesgo pais",
                        "excel_equiv": "=Rf+Beta*(Rm-Rf)+RiesgoPais",
                    },
                    {
                        "type": "practice",
                        "worksheet": "capm_rates",
                        "title": "Calcula la tasa de actualizacion",
                        "source_sheet": "Tasas",
                    },
                ],
            },
            {
                "id": "m4-van-tir",
                "order": 4,
                "title": "VAN y TIR (evaluacion economica)",
                "duration_min": 40,
                "source": [
                    "Capitulo VII. Evaluacion de proyectos.pptx",
                    "S7 Ejercicios de evaluacion de proyectos.xlsx",
                ],
                "objectives": [
                    "Calcular VAN = NPV(tasa, flujos)",
                    "Interpretar TIR vs costo de oportunidad",
                ],
                "tips": [
                    "VAN > 0 => proyecto crea valor",
                    "TIR > COK => rentable en terminos relativos",
                    "Regla del 5%: si VAN difieren menos de 5%, son equivalentes",
                ],
                "steps": [
                    {
                        "type": "formula",
                        "title": "VAN en Excel",
                        "formula": "VAN = NPV(tasa, flujos_periodo_1..n) + Flujo_periodo_0",
                        "excel_equiv": "=NPV(B5,C24:L24)+C23",
                    },
                    {
                        "type": "formula",
                        "title": "TIR en Excel",
                        "formula": "TIR = IRR(rango_flujos_incluyendo_inversion_inicial)",
                        "excel_equiv": "=IRR(B24:V24)",
                    },
                    {
                        "type": "practice",
                        "worksheet": "van_tir",
                        "title": "Evalua un proyecto simple",
                    },
                ],
            },
            {
                "id": "m5-financiamiento",
                "order": 5,
                "title": "Financiamiento y FCF",
                "duration_min": 50,
                "source": [
                    "S8 Ejerc icios financiamiento.xls",
                    "S6 Flujo basico FCE y FCF.xlsx",
                ],
                "objectives": [
                    "Construir tabla de amortizacion (capital, interes, cuota)",
                    "Diferenciar VANE/TIRE vs VANF/TIRF",
                ],
                "tips": [
                    "Cuota constante: PMT(tasa, plazo, monto_prestamo)",
                    "Interes periodo = Saldo anterior * tasa",
                    "Amortizacion = Cuota - Interes",
                    "Con deuda: UAI = UAII - Intereses (escudo fiscal)",
                ],
                "steps": [
                    {
                        "type": "formula",
                        "title": "Cuota constante",
                        "formula": "Cuota = PMT(i, n, -Prestamo)",
                        "excel_equiv": "=-PMT($B$5,$B$8,B15)",
                    },
                    {
                        "type": "practice",
                        "worksheet": "amortization",
                        "title": "Tabla de amortizacion 3 periodos",
                    },
                    {
                        "type": "practice",
                        "worksheet": "fcf_financing",
                        "title": "FCE vs FCF con prestamo de 500",
                    },
                ],
            },
            {
                "id": "m6-escenarios",
                "order": 6,
                "title": "Escenarios y decision",
                "duration_min": 40,
                "source": [
                    "S8 Ejerc icios financiamiento.xls",
                    "Aplicaciones flujo de caja.xlsx",
                ],
                "objectives": [
                    "Comparar escenarios con/sin deuda",
                    "Recomendar alternativa segun VAN y rentabilidad",
                ],
                "tips": [
                    "Mas deuda puede subir VANF pero tambien el riesgo",
                    "Compara siempre a misma tasa de descuento",
                ],
                "steps": [
                    {
                        "type": "practice",
                        "worksheet": "scenarios",
                        "title": "Caso base vs prestamo 500 vs ingresos menores",
                    },
                ],
            },
            {
                "id": "m7-integrador",
                "order": 7,
                "title": "Caso integrador (simulacro examen)",
                "duration_min": 60,
                "source": [
                    "Ejercicios de evaluacion de proyectos.docx",
                    "_BEG06M-Tercera PC 2023_II-solucionario.xlsx",
                ],
                "objectives": [
                    "Resolver caso completo: flujos + financiamiento + VAN + recomendacion",
                ],
                "steps": [
                    {
                        "type": "practice",
                        "worksheet": "exam_simulation",
                        "title": "Proyecto eolico simplificado",
                    },
                ],
            },
        ],
    }


def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    results = []
    for fp in sorted(BASE.glob("*.xlsx")):
        try:
            results.append(analyze_xlsx(fp))
            print(f"OK {fp.name}")
        except Exception as exc:
            print(f"ERR {fp.name}: {exc}")

    analysis_path = OUT
    analysis_path.write_text(
        json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    curriculum = build_curriculum(results)
    curriculum_path = OUT.parent / "curriculum.json"
    curriculum_path.write_text(
        json.dumps(curriculum, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"Wrote {analysis_path}")
    print(f"Wrote {curriculum_path}")


if __name__ == "__main__":
    main()
