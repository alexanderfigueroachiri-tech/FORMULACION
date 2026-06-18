"""Extract exact formulas from key course Excel sheets."""
import json
from pathlib import Path

import openpyxl
import xlrd

BASE = Path(__file__).resolve().parent.parent / "BASE"
OUT = Path(__file__).resolve().parent.parent / "data" / "excel_formulas_exact.json"


def extract_xlsx(path: Path, sheets: list[str]) -> dict:
    wb = openpyxl.load_workbook(path, data_only=False)
    result = {}
    for sn in sheets:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        cells = {}
        for row in ws.iter_rows(max_row=35, max_col=16):
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                label = ws.cell(cell.row, 2).value if cell.column > 2 else None
                cells[cell.coordinate] = {
                    "row_label": str(label).strip() if label else "",
                    "value": str(v),
                    "is_formula": isinstance(v, str) and v.startswith("="),
                }
        result[sn] = cells
    wb.close()
    return result


def main():
    data = {}
    s6 = BASE / "S6 Flujo basico FCE y FCF.xlsx"
    s7 = BASE / "S7 Ejercicios de  evaluación de proyectos.xlsx"
    if s6.exists():
        data["S6"] = extract_xlsx(s6, ["FCE", "FCEF", "Tasas"])
    if s7.exists():
        data["S7"] = extract_xlsx(s7, ["Nuevo o usado", "Eolico", "Vidas_dife"])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
