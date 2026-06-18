import json
from pathlib import Path

import openpyxl

base = Path(r"C:\Users\alexd\Projects\formulacion\BASE")
results = {}

for fp in sorted(base.glob("*.xlsx")):
    try:
        wb = openpyxl.load_workbook(fp, data_only=False)
        entry = {"sheets": wb.sheetnames, "formulas_by_sheet": {}, "formula_counts": {}}
        for sn in wb.sheetnames:
            ws = wb[sn]
            formulas = []
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        formulas.append({"cell": cell.coordinate, "formula": value})
            if formulas:
                entry["formulas_by_sheet"][sn] = formulas[:120]
                entry["formula_counts"][sn] = len(formulas)
        wb.close()
        results[fp.name] = entry
    except Exception as exc:
        results[fp.name] = {"error": str(exc)}

out = Path(r"C:\Users\alexd\Projects\formulacion\formulas_extracted.json")
out.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Wrote {out}")
for name, data in results.items():
    if "error" in data:
        print(f"{name}: ERROR {data['error']}")
    else:
        counts = data.get("formula_counts", {})
        total = sum(counts.values())
        print(f"{name}: {total} formulas across {len(counts)} sheets")
